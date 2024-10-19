import openpyxl
import json
import re


def extract_formulas(file_path):
    # Load the workbook and select all the sheets
    workbook = openpyxl.load_workbook(file_path, data_only=False)

    # Dictionary to store formulas from all sheets
    formulas = {}

    # Iterate through each sheet in the workbook
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        column_headers = {
            cell.column_letter: cell.value for cell in sheet[1]
        }  # Assumes headers are in the first row
        sheet_formulas = {}

        # Iterate through each cell in the sheet, starting from the second row to skip headers
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if cell.data_type == "f":  # Check if the cell contains a formula
                    # Get the column header or fallback to the cell coordinate
                    column_header = column_headers.get(
                        cell.column_letter, cell.column_letter
                    )
                    # Use column header and row number as the key (e.g., "Price_7")
                    key = (
                        f"{column_header}_{cell.row}"
                        if column_header
                        else cell.coordinate
                    )
                    sheet_formulas[key] = cell.value

        # Add formulas from the sheet to the overall dictionary
        if sheet_formulas:
            formulas[sheet_name] = sheet_formulas

    return formulas


def normalize_formula(formula):
    """
    Normalizes the formula by replacing cell references (e.g., A1, $B$2) with a placeholder.
    This allows formulas with the same structure but different cell references to be considered identical.
    """
    cell_reference_pattern = r"(\$?[A-Z]+\$?\d+)"
    normalized_formula = re.sub(cell_reference_pattern, "CELL_REF", formula)
    return normalized_formula


def extract_unique_formulas(formulas_dict):
    unique_formulas_set = set()  # To store unique normalized formulas
    unique_formulas = (
        {}
    )  # To store unique formulas with their corresponding original formulas

    for sheet_name, sheet_formulas in formulas_dict.items():
        for cell, formula in sheet_formulas.items():
            # Normalize the formula to remove specific cell references
            normalized_formula = normalize_formula(formula)

            # If the normalized formula is not in the set, it's unique
            if normalized_formula not in unique_formulas_set:
                unique_formulas_set.add(normalized_formula)
                unique_formulas[f"{sheet_name}_{cell}"] = formula

    return unique_formulas


def save_formulas_to_json_file(formulas, output_file):
    with open(output_file, "w") as f:
        json.dump(formulas, f, indent=4)


# Usage
file_path = "BM055unprotected.xlsx"

# Extract all formulas and save to "extracted_formulas.json"
formulas = extract_formulas(file_path)
save_formulas_to_json_file(formulas, "extracted_formulas.json")

# Extract unique formulas and save to "unique_formulas.json"
unique_formulas = extract_unique_formulas(formulas)
save_formulas_to_json_file(unique_formulas, "unique_formulas.json")
